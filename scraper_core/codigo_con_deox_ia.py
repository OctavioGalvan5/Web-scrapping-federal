# -*- coding: utf-8 -*-
"""
Scraper de Expedientes con Análisis IA - Portal PJN
Integra extracción de expedientes y análisis individual con IA
VERSIÓN CORREGIDA - Problemas de notificaciones y paginación solucionados
MODIFICACIÓN: Solo se cambió la función de análisis para extraer datos HTML en lugar de PDFs
"""

import requests
import pdfplumber
import google.generativeai as genai
import json
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from PIL import Image
import pytesseract
import time
import pandas as pd
import os
from openpyxl import load_workbook
import traceback
from bs4 import BeautifulSoup
from datetime import datetime

# ==================== CONFIGURACIÓN INICIAL ====================

def configurar_tesseract():
    """Configura Tesseract OCR"""
    try:
        tesseract_paths = [
            r'C:\Users\octav\AppData\Local\Programs\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            '/usr/bin/tesseract',
            '/usr/local/bin/tesseract'
        ]
        
        for path in tesseract_paths:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                print(f"✅ Tesseract configurado: {path}")
                return True
        
        print("⚠️ Tesseract no encontrado en rutas comunes")
        return False
        
    except Exception as e:
        print(f"⚠️ Error configurando Tesseract: {e}")
        return False

# ==================== FUNCIONES DE EXTRACCIÓN (CÓDIGO ORIGINAL) ====================

def extraer_numero_y_ano_expediente(expediente):
    """Extrae el número y año de un expediente - VERSIÓN MEJORADA"""
    try:
        if not expediente or expediente == "No disponible":
            return "", ""
        
        expediente_limpio = expediente.strip()
        print(f"   🔍 Analizando expediente: '{expediente_limpio}'")
        
        # PATRÓN 1: Con prefijo (FSA, etc.)
        patron_con_prefijo = r'([A-Z]{2,5})\s*(\d{1,9})/(\d{4})(?:/.*)?'
        match = re.search(patron_con_prefijo, expediente_limpio)
        
        if match:
            prefijo = match.group(1)
            numero = match.group(2)
            ano = match.group(3)
            
            print(f"       ✅ PATRÓN CON PREFIJO detectado:")
            print(f"           → Prefijo: {prefijo}")
            print(f"           → Número: {numero}")
            print(f"           → Año: {ano}")
            
            return numero, ano
        
        # PATRÓN 2: Sin prefijo
        patron_sin_prefijo = r'^(\d{1,9})/(\d{4})(?:/.*)?$'
        match = re.search(patron_sin_prefijo, expediente_limpio)
        
        if match:
            numero = match.group(1)
            ano = match.group(2)
            
            print(f"       ✅ PATRÓN SIN PREFIJO detectado:")
            print(f"           → Número: {numero}")
            print(f"           → Año: {ano}")
            
            return numero, ano
        
        # PATRÓN 3: Flexible
        patron_flexible = r'(\d{1,9})/(\d{4})'
        match = re.search(patron_flexible, expediente_limpio)
        
        if match:
            numero = match.group(1)
            ano = match.group(2)
            
            print(f"       ✅ PATRÓN FLEXIBLE detectado:")
            print(f"           → Número: {numero}")
            print(f"           → Año: {ano}")
            
            return numero, ano
        
        print(f"       ❌ No se pudo extraer con ningún patrón")
        return "", ""
        
    except Exception as e:
        print(f"   ❌ Error extrayendo número y año de '{expediente}': {str(e)}")
        return "", ""

def normalizar_fecha_para_comparacion(fecha):
    """Normaliza una fecha quitando ceros iniciales para comparación"""
    try:
        if '/' in fecha:
            partes = fecha.split('/')
            if len(partes) == 3:
                dia = str(int(partes[0]))
                mes = str(int(partes[1]))
                ano = partes[2]
                return f"{dia}/{mes}/{ano}"
        return fecha
    except (ValueError, IndexError) as e:
        print(f"   ⚠️ Error normalizando fecha '{fecha}': {str(e)}")
        return fecha

def normalizar_fecha_para_excel(fecha):
    """Normaliza una fecha CON ceros iniciales para Excel"""
    try:
        if '/' in fecha:
            partes = fecha.split('/')
            if len(partes) == 3:
                dia = partes[0].zfill(2)
                mes = partes[1].zfill(2)
                ano = partes[2]
                return f"{dia}/{mes}/{ano}"
        return fecha
    except Exception as e:
        print(f"   ⚠️ Error normalizando fecha para Excel '{fecha}': {str(e)}")
        return fecha

def comparar_fechas_mejorado(fecha1, fecha2):
    """Compara dos fechas normalizando ambas"""
    try:
        fecha_norm1 = normalizar_fecha_para_comparacion(fecha1)
        fecha_norm2 = normalizar_fecha_para_comparacion(fecha2)
        return fecha_norm1 == fecha_norm2
    except Exception as e:
        print(f"   ⚠️ Error comparando fechas: {str(e)}")
        return False

# ==================== FUNCIONES DE ANÁLISIS IA ====================

def solve_recaptcha(api_key, sitekey, pageurl):
    """Resuelve reCAPTCHA usando 2captcha"""
    if not api_key or api_key == "TU_CLAVE_2CAPTCHA":
        print("Error: API Key de 2Captcha no proporcionada.")
        return None
    
    s = requests.Session()
    captcha_url = "http://2captcha.com/in.php"
    data = {
        "key": api_key,
        "method": "userrecaptcha",
        "googlekey": sitekey,
        "pageurl": pageurl,
        "json": 1
    }
    
    try:
        response = s.post(captcha_url, data=data, timeout=30)
        response.raise_for_status()
        result = response.json()
        
        if result.get("status") != 1:
            print(f"Error al enviar CAPTCHA a 2captcha: {result.get('request')}")
            return None
        
        captcha_id = result.get("request")
        print(f"Captcha enviado a 2captcha, ID: {captcha_id}. Esperando solución...")
        
        fetch_url = "http://2captcha.com/res.php"
        params = {"key": api_key, "action": "get", "id": captcha_id, "json": 1}
        start_time = time.time()
        timeout_2captcha = 180
        
        while (time.time() - start_time) < timeout_2captcha:
            time.sleep(10)
            res = s.get(fetch_url, params=params, timeout=30)
            res.raise_for_status()
            result2 = res.json()
            
            if result2.get("status") == 1:
                solution = result2.get("request")
                print(f"Captcha resuelto por 2captcha: {solution[:15]}...")
                return solution
            elif result2.get("request") == "CAPCHA_NOT_READY":
                print("Captcha aún no está listo...")
                continue
            else:
                print(f"Error obteniendo solución de 2captcha: {result2.get('request')}")
                return None
        
        print("Timeout esperando la solución del CAPTCHA.")
        return None
        
    except requests.exceptions.RequestException as e:
        print(f"Error de conexión con 2captcha: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"Error decodificando respuesta JSON de 2captcha: {e}")
        return None

# ==================== NUEVAS FUNCIONES PARA EXTRACCIÓN HTML ====================

def extraer_datos_basicos_expediente(driver):
    """
    NUEVA FUNCIÓN: Extrae los datos básicos del expediente del HTML
    """
    datos_expediente = {
        'expediente': 'No disponible',
        'jurisdiccion': 'No disponible', 
        'dependencia': 'No disponible',
        'situacion_actual': 'No disponible',
        'caratula': 'No disponible'
    }
    
    try:
        print("📄 Extrayendo datos básicos del expediente...")
        
        # Esperar a que cargue el contenido principal
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "ui-fieldset-content"))
        )
        
        # Obtener el HTML de la página
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Buscar el contenedor principal
        fieldset_content = soup.find('div', class_='ui-fieldset-content')
        
        if fieldset_content:
            # Extraer cada campo usando los labels
            labels_map = {
                'Expediente:': 'expediente',
                'Jurisdicción:': 'jurisdiccion', 
                'Dependencia:': 'dependencia',
                'Sit. Actual:': 'situacion_actual',
                'Carátula:': 'caratula'
            }
            
            for label_text, key in labels_map.items():
                try:
                    # Buscar el label
                    label = fieldset_content.find('label', string=label_text)
                    if label:
                        # Encontrar el contenedor padre y buscar el span con el valor
                        form_group = label.find_parent('div', class_='form-group')
                        if form_group:
                            # Buscar el span con el valor (puede tener diferentes clases)
                            value_span = form_group.find('span', style=lambda x: x and 'color:#000000' in x)
                            if not value_span:
                                value_span = form_group.find('span', class_='form_control')
                            
                            if value_span:
                                datos_expediente[key] = value_span.get_text(strip=True)
                                print(f"   ✅ {label_text} {datos_expediente[key]}")
                            else:
                                print(f"   ⚠️ No se encontró valor para {label_text}")
                        else:
                            print(f"   ⚠️ No se encontró form-group para {label_text}")
                    else:
                        print(f"   ⚠️ No se encontró label {label_text}")
                        
                except Exception as e:
                    print(f"   ❌ Error extrayendo {label_text}: {e}")
                    continue
        else:
            print("   ❌ No se encontró el contenedor ui-fieldset-content")
            
    except Exception as e:
        print(f"❌ Error general extrayendo datos básicos: {e}")
    
    return datos_expediente

def hacer_click_intervinientes_y_extraer(driver):
    """
    FUNCIÓN NUEVA CON IA: Hace clic en Intervinientes y usa IA para extraer el letrado principal
    """
    datos_intervinientes = {
        'actor_nombre': 'No disponible',
        'letrado_apoderado': 'No disponible',
        'tomo_folio': 'No disponible',
        'cuit_cuil': 'No disponible'
    }
    
    try:
        print("👥 Haciendo clic en pestaña Intervinientes...")
        
        # Buscar y hacer clic en la pestaña Intervinientes con múltiples intentos
        tab_found = False
        max_intentos = 3
        
        for intento in range(max_intentos):
            try:
                print(f"   🔄 Intento {intento + 1}/{max_intentos} para encontrar pestaña Intervinientes...")
                
                # Esperar más tiempo y usar múltiples selectores
                selectores_tab = [
                    "//span[contains(@class, 'rf-tab-lbl') and text()='Intervinientes']",
                    "//span[text()='Intervinientes']",
                    "//a[contains(@class, 'rf-tab-lbl') and text()='Intervinientes']"
                ]
                
                for selector in selectores_tab:
                    try:
                        tab_intervinientes = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        driver.execute_script("arguments[0].click();", tab_intervinientes)
                        print(f"   ✅ Clic en Intervinientes realizado con selector: {selector}")
                        tab_found = True
                        break
                    except:
                        continue
                
                if tab_found:
                    break
                    
                # Si no encontró la pestaña, esperar un poco más
                time.sleep(2)
                
            except Exception as e:
                print(f"   ⚠️ Intento {intento + 1} falló: {e}")
                if intento < max_intentos - 1:
                    time.sleep(3)
        
        if not tab_found:
            print("   ❌ No se pudo encontrar la pestaña Intervinientes después de múltiples intentos")
            return datos_intervinientes
        
        # Esperar a que cargue la tabla específica de participantes con más tiempo
        print("   ⏳ Esperando carga de tabla de participantes...")
        time.sleep(8)  # Más tiempo de espera
        
        # Intentar encontrar la tabla con múltiples estrategias
        tabla_encontrada = False
        for intento_tabla in range(3):
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "expediente:participantsTable"))
                )
                tabla_encontrada = True
                break
            except:
                print(f"   ⚠️ Intento {intento_tabla + 1} para cargar tabla...")
                time.sleep(5)
        
        if not tabla_encontrada:
            print("   ❌ No se pudo cargar la tabla de participantes")
            return datos_intervinientes
        
        print("   📋 Tabla de participantes cargada, extrayendo HTML completo...")
        
        # ==================== NUEVA ESTRATEGIA: EXTRAER HTML Y USAR IA ====================
        
        # Obtener el HTML de la tabla completa
        try:
            tabla_element = driver.find_element(By.ID, "expediente:participantsTable")
            html_tabla = tabla_element.get_attribute('outerHTML')
            
            print("   🤖 Enviando HTML de la tabla a la IA para análisis...")
            
            # Usar IA para extraer la información
            datos_ia = extraer_letrado_con_ia(html_tabla)
            
            if datos_ia:
                datos_intervinientes.update(datos_ia)
                print("   ✅ Datos extraídos exitosamente por la IA")
            else:
                print("   ⚠️ La IA no pudo extraer los datos")
                
        except Exception as e:
            print(f"   ❌ Error obteniendo HTML de la tabla: {e}")
                    
    except TimeoutException:
        print("   ❌ Timeout: No se pudo encontrar la pestaña Intervinientes")
    except Exception as e:
        print(f"   ❌ Error general en intervinientes: {e}")
        import traceback
        traceback.print_exc()
    
    return datos_intervinientes

def extraer_letrado_con_ia(html_tabla, gemini_api_key=None):
    """
    NUEVA FUNCIÓN: Usa IA para extraer el letrado apoderado principal del HTML de la tabla
    """
    try:
        # Configurar Gemini
        genai.configure(api_key=gemini_api_key)
        
        prompt = f"""**Rol:** Eres un experto en análisis de expedientes judiciales argentinos.

**Tarea:** Analizar la siguiente tabla HTML de participantes de un expediente y extraer información específica.

**Instrucciones IMPORTANTES:**
1. Busca TODOS los participantes con tipo "ACTOR"
2. Para cada ACTOR, busca su "LETRADO APODERADO" (no otros tipos de letrados)
3. Si hay múltiples actores con letrados, elige el PRINCIPAL (generalmente el primero o el más relevante)
4. Si hay múltiples letrados apoderados para el mismo actor, elige UNO solo (el primero)
5. RESPONDE ÚNICAMENTE CON UN JSON VÁLIDO, SIN TEXTO ADICIONAL

**Estructura JSON requerida:**
{{
    "actor_nombre": "nombre completo del actor principal",
    "letrado_apoderado": "nombre completo del letrado apoderado principal",
    "tomo_folio": "información de tomo y folio del letrado",
    "cuit_cuil": "CUIT/CUIL del letrado"
}}

**Si no encuentras información, usa "No disponible" para ese campo.**

**HTML de la tabla a analizar:**
{html_tabla}
"""
        
        print("   🤖 Enviando consulta a Gemini...")
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt, request_options={'timeout': 60})
        
        if hasattr(response, 'text'):
            respuesta_texto = response.text.strip()
            print("   ✅ Respuesta recibida de Gemini")
            
            # Buscar el JSON en la respuesta
            try:
                inicio_json = respuesta_texto.find('{')
                fin_json = respuesta_texto.rfind('}') + 1
                
                if inicio_json != -1 and fin_json != -1:
                    json_str = respuesta_texto[inicio_json:fin_json]
                    datos_json = json.loads(json_str)
                    
                    print(f"   📋 Datos extraídos por IA:")
                    print(f"      Actor: {datos_json.get('actor_nombre', 'No disponible')}")
                    print(f"      Letrado: {datos_json.get('letrado_apoderado', 'No disponible')}")
                    print(f"      Tomo/Folio: {datos_json.get('tomo_folio', 'No disponible')}")
                    print(f"      CUIT: {datos_json.get('cuit_cuil', 'No disponible')}")
                    
                    return datos_json
                else:
                    print("   ❌ No se encontró JSON válido en la respuesta")
                    return None
                    
            except json.JSONDecodeError as e:
                print(f"   ❌ Error decodificando JSON: {e}")
                print(f"   📄 Respuesta completa: {respuesta_texto[:500]}...")
                return None
        else:
            print("   ❌ La respuesta de Gemini no tiene texto")
            return None
            
    except Exception as e:
        print(f"   ❌ Error llamando a la API de Gemini: {e}")
        return None

def analyze_legal_documents_html(datos_expediente, gemini_api_key):
    """
    FUNCIÓN MODIFICADA: Analiza datos de expediente extraídos del HTML usando Gemini - DEVUELVE JSON
    """
    if not gemini_api_key:
        print("Error: API Key de Gemini no configurada.")
        return ""
    
    if not datos_expediente:
        print("Advertencia: No hay datos del expediente para analizar.")
        return ""
    
    try:
        # Configurar Gemini
        genai.configure(api_key=gemini_api_key)
        
        # Crear el texto con la información extraída
        texto_analisis = f"""
INFORMACIÓN DEL EXPEDIENTE:
- Expediente: {datos_expediente.get('expediente', 'No disponible')}
- Jurisdicción: {datos_expediente.get('jurisdiccion', 'No disponible')}
- Dependencia: {datos_expediente.get('dependencia', 'No disponible')}
- Situación Actual: {datos_expediente.get('situacion_actual', 'No disponible')}
- Carátula: {datos_expediente.get('caratula', 'No disponible')}

INTERVINIENTES:
- Actor: {datos_expediente.get('actor_nombre', 'No disponible')}
- Letrado Apoderado: {datos_expediente.get('letrado_apoderado', 'No disponible')}
- Tomo/Folio: {datos_expediente.get('tomo_folio', 'No disponible')}
- CUIT/CUIL: {datos_expediente.get('cuit_cuil', 'No disponible')}
"""
        
        prompt = f"""**Rol:** Eres un asistente legal experto especializado en análisis de expedientes judiciales argentinos.

**Tarea:** Analizar y limpiar los siguientes datos de un expediente judicial, devolviendo ÚNICAMENTE un JSON válido.

**Instrucciones IMPORTANTES:**
1. Limpia y normaliza todos los datos
2. Extrae información relevante de la carátula (tipo de proceso, partes, etc.)
3. Identifica el tipo de juzgado y jurisdicción
4. Normaliza nombres y datos personales
5. RESPONDE ÚNICAMENTE CON UN JSON VÁLIDO, SIN TEXTO ADICIONAL
6. NO agregues explicaciones, comentarios o texto fuera del JSON

**Estructura JSON requerida:**
{{
    "expediente_numero": "número limpio del expediente",
    "jurisdiccion_normalizada": "jurisdicción limpia",
    "juzgado": "nombre del juzgado normalizado",
    "secretaria": "secretaría si está disponible",
    "situacion_actual": "situación normalizada",
    "tipo_proceso": "tipo de proceso extraído de la carátula",
    "actor_principal": "nombre del actor principal limpio",
    "demandado_principal": "demandado extraído de la carátula",
    "objeto_demanda": "objeto de la demanda",
    "letrado_actor": "nombre del letrado limpio",
    "matricula_letrado": "tomo y folio del letrado",
    "cuit_cuil_letrado": "CUIT/CUIL del letrado limpio",
    "observaciones": "cualquier observación relevante"
}}

**Información del Expediente a Analizar:**
{texto_analisis}
"""
        
        print("🤖 Enviando información a Gemini para análisis JSON...")
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt, request_options={'timeout': 600})
        
        if hasattr(response, 'text'):
            print("✅ Análisis JSON completado por Gemini")
            return response.text.strip()
        else:
            print("Error: La respuesta de Gemini no tiene texto.")
            return ""
            
    except Exception as e:
        print(f"Error llamando a la API de Gemini: {e}")
        return ""

# ==================== FUNCIONES ORIGINALES (SIN CAMBIOS) ====================

def descargar_y_extraer_pdf(url_pdf, textos):
    """Descarga y extrae texto de un PDF con OCR mejorado"""
    nombre_pdf = "temp_document.pdf"
    try:
        print(f"📄 Descargando PDF: {url_pdf}")
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        resp = requests.get(url_pdf, headers=headers, timeout=60, stream=True)
        resp.raise_for_status()
        
        with open(nombre_pdf, 'wb') as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        
        print(f"✅ PDF descargado exitosamente")
        
        texto_completo_pdf = ''
        paginas_con_ocr = 0
        paginas_sin_texto = 0
        
        try:
            with pdfplumber.open(nombre_pdf) as pdf:
                total_paginas = len(pdf.pages)
                print(f"📖 Procesando PDF con {total_paginas} páginas...")
                
                for i, page in enumerate(pdf.pages):
                    print(f"  📄 Procesando página {i+1}/{total_paginas}...")
                    
                    # Intentar extraer texto normal
                    page_text = page.extract_text()
                    
                    if page_text and len(page_text.strip()) > 20:
                        # Texto extraído correctamente
                        texto_completo_pdf += page_text + "\n"
                        print(f"    ✅ Texto extraído normalmente ({len(page_text.strip())} caracteres)")
                    else:
                        # Página sin texto o con muy poco texto - aplicar OCR
                        paginas_sin_texto += 1
                        print(f"    ⚠️ Página con poco/sin texto ({len(page_text.strip()) if page_text else 0} chars)")
                        
                        # Verificar si Tesseract está disponible
                        if pytesseract.pytesseract.tesseract_cmd and os.path.exists(pytesseract.pytesseract.tesseract_cmd):
                            print(f"    🔍 Aplicando OCR con Tesseract...")
                            try:
                                # Convertir página a imagen con alta resolución
                                pil_image = page.to_image(resolution=300).original
                                
                                # Aplicar OCR
                                ocr_text = pytesseract.image_to_string(
                                    pil_image, 
                                    lang="spa",
                                    config='--psm 6 --oem 3'  # Configuración optimizada
                                )
                                
                                if ocr_text and len(ocr_text.strip()) > 10:
                                    texto_completo_pdf += ocr_text + "\n"
                                    paginas_con_ocr += 1
                                    print(f"    ✅ OCR exitoso ({len(ocr_text.strip())} caracteres extraídos)")
                                else:
                                    print(f"    ⚠️ OCR no extrajo texto útil")
                                    
                            except Exception as ocr_e:
                                print(f"    ❌ Error durante OCR: {ocr_e}")
                        else:
                            print(f"    ❌ Tesseract no disponible para OCR")
                            if not pytesseract.pytesseract.tesseract_cmd:
                                print(f"    💡 Instale Tesseract OCR para extraer texto de imágenes")
            
            # Resumen del procesamiento
            print(f"📊 Resumen del PDF:")
            print(f"    - Total páginas: {total_paginas}")
            print(f"    - Páginas con OCR aplicado: {paginas_con_ocr}")
            print(f"    - Páginas sin texto: {paginas_sin_texto}")
            print(f"    - Texto total extraído: {len(texto_completo_pdf)} caracteres")
            
            if texto_completo_pdf.strip():
                textos.append(texto_completo_pdf)
                print(f"✅ PDF procesado y texto agregado")
            else:
                print(f"⚠️ PDF no contenía texto extraíble")
                
        except Exception as pdf_e:
            print(f"❌ Error procesando el PDF '{nombre_pdf}': {pdf_e}")
            
    except requests.exceptions.RequestException as req_e:
        print(f"❌ Error descargando PDF desde {url_pdf}: {req_e}")
    except Exception as e:
        print(f"❌ Error general en descargar_y_extraer_pdf para {url_pdf}: {e}")
    finally:
        if os.path.exists(nombre_pdf):
            try:
                os.remove(nombre_pdf)
                print(f"🗑️ Archivo temporal eliminado")
            except Exception as del_e:
                print(f"⚠️ Error eliminando archivo temporal '{nombre_pdf}': {del_e}")

def extraer_info_pagina(driver, tabla_id, textos):
    """Extrae información de una página/tabla"""
    try:
        tabla = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, tabla_id))
        )
        filas = tabla.find_elements(By.TAG_NAME, "tr")
        print(f"Procesando tabla: {tabla_id}. Filas encontradas: {len(filas)}")
        
        for i, fila in enumerate(filas):
            celdas = fila.find_elements(By.TAG_NAME, "td")
            if celdas:
                enlaces = fila.find_elements(By.TAG_NAME, "a")
                for enlace in enlaces:
                    href = enlace.get_attribute("href")
                    if href and ("download=true" in href or ".pdf" in href.lower()):
                        print(f"    Encontrado enlace PDF: {href}")
                        descargar_y_extraer_pdf(href, textos)
                        
    except Exception as e:
        print(f"Error extrayendo datos de la tabla '{tabla_id}': {e}")

def analyze_legal_documents(texto_pdfs, gemini_api_key):
    """Analiza documentos legales usando Gemini"""
    if not gemini_api_key:
        print("Error: API Key de Gemini no configurada.")
        return ""
    
    if not texto_pdfs or not texto_pdfs.strip():
        print("Advertencia: No hay texto extraído de PDFs para analizar.")
        return ""
    
    try:
        # Configurar Gemini
        genai.configure(api_key=gemini_api_key)
        
        prompt = (
            "**Rol:** Eres un asistente legal experto, especializado en el análisis detallado de expedientes judiciales argentinos.\n"
            "**Tarea:** Analizar el texto de un expediente judicial y proporcionar un resumen completo del caso.\n\n"
            "**Instrucciones:**\n"
            "1. **Resumen del Caso:** Proporciona un resumen claro y conciso del caso, incluyendo:\n"
            "   - Tipo de proceso judicial\n"
            "   - Partes involucradas (actor/demandado)\n"
            "   - Objeto de la demanda o conflicto principal\n"
            "   - Estado actual del proceso\n"
            "   - Resoluciones importantes\n\n"
            "2. **Información sobre Aportes y Estampillas:**\n"
            "   - Determina si se pagaron los aportes requeridos\n"
            "   - Identifica los abogados intervinientes\n"
            "   - Verifica el pago de estampillas por cada abogado\n"
            "   - Lista abogados con estampillas pendientes\n\n"
            "**Formato de Salida:**\n"
            "Estructura tu respuesta en secciones claras con subtítulos.\n"
            "Si alguna información no está disponible, indícalo explícitamente.\n\n"
            f"**Texto del Expediente a Analizar:**\n{texto_pdfs}\n"
        )
        
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt, request_options={'timeout': 600})
        
        if hasattr(response, 'text'):
            return response.text
        else:
            print("Error: La respuesta de Gemini no tiene texto.")
            return ""
            
    except Exception as e:
        print(f"Error llamando a la API de Gemini: {e}")
        return ""

# ==================== FUNCIONES CORREGIDAS ====================

def ir_siguiente_pagina_notificaciones(driver):
    """
    NUEVA FUNCIÓN: Navega a la siguiente página de notificaciones
    """
    try:
        print("   🔄 Buscando botón 'Siguiente' en notificaciones...")
        
        # Estrategias para encontrar el botón siguiente en Material-UI
        selectores_siguiente = [
            "//button[@aria-label='Go to next page']",
            "//button[@title='Go to next page']",
            "//button[contains(@class, 'MuiIconButton-root') and contains(@aria-label, 'next')]",
            "//button[contains(@class, 'MuiIconButton-root') and not(@disabled)]//svg[contains(@data-testid, 'KeyboardArrowRight')]",
            "//div[contains(@class, 'MuiTablePagination-actions')]//button[last()]",
            "//button[contains(@aria-label, 'siguiente')]",
            "//button[contains(@aria-label, 'next')]"
        ]
        
        for selector in selectores_siguiente:
            try:
                boton_siguiente = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, selector))
                )
                
                # Verificar que no esté deshabilitado
                if boton_siguiente.get_attribute("disabled") or "disabled" in boton_siguiente.get_attribute("class"):
                    print(f"   ⚠️ Botón encontrado pero está deshabilitado: {selector}")
                    continue
                
                # Hacer scroll al botón y hacer clic
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_siguiente)
                time.sleep(0.5)
                
                # Intentar clic normal primero
                try:
                    boton_siguiente.click()
                except:
                    # Si falla, usar JavaScript
                    driver.execute_script("arguments[0].click();", boton_siguiente)
                
                print(f"   ✅ Click exitoso en botón siguiente: {selector}")
                return True
                
            except TimeoutException:
                continue
            except Exception as e:
                print(f"   ⚠️ Error con selector {selector}: {e}")
                continue
        
        print("   ❌ No se encontró botón 'Siguiente' funcional")
        return False
        
    except Exception as e:
        print(f"   ❌ Error general navegando a siguiente página: {e}")
        return False

XPATH_FILAS_NOTIF = "//tr[contains(@class, 'MuiBox-root') and @role='row']"

def _esperar_filas_estables(driver, segundos_estable=2, timeout=25):
    """Espera hasta que el conteo de filas no cambie durante `segundos_estable` segundos."""
    prev = -1
    estable = 0
    inicio = time.time()
    while time.time() - inicio < timeout:
        count = len(driver.find_elements(By.XPATH, XPATH_FILAS_NOTIF))
        if count > 1 and count == prev:
            estable += 1
            if estable >= segundos_estable:
                return count
        else:
            estable = 0
        prev = count
        time.sleep(1)
    return prev


def procesar_filas_notificaciones_con_paginacion(driver, fecha_objetivo, filas_consultas):
    """
    FUNCIÓN CORREGIDA: Procesa las notificaciones con paginación mejorada
    """
    filas_notificaciones = []
    expedientes_procesados = set(f['expediente'].strip().upper() for f in filas_consultas)

    pagina_actual = 1

    while True:
        print(f"\n📄 === PROCESANDO PÁGINA {pagina_actual} DE NOTIFICACIONES ===")

        try:
            # Esperar a que el conteo de filas se estabilice (no solo a que haya >1)
            count = _esperar_filas_estables(driver)
            print(f"   ⏳ Tabla estable con {count} filas")

            filas = driver.find_elements(By.XPATH, XPATH_FILAS_NOTIF)
            print(f"📋 Encontradas {len(filas)} filas de notificaciones en página {pagina_actual}")
            
            filas_pagina_actual = 0
            
            for i, fila in enumerate(filas):
                try:
                    # Extraer fecha del aria-label
                    aria_label = fila.get_attribute("aria-label")
                    fecha_extraida = None
                    
                    if aria_label:
                        match = re.search(r'Fecha:\s*(\d{1,2}/\d{1,2}/\d{4})', aria_label)
                        if match:
                            fecha_extraida = match.group(1)
                    
                    if not fecha_extraida:
                        continue
                    
                    # Comparar fechas (usando tu función existente)
                    if comparar_fechas_mejorado(fecha_extraida, fecha_objetivo):
                        print(f"   ✅ COINCIDENCIA DE FECHA en fila {i+1}:")
                        print(f"       Extraída: '{fecha_extraida}' vs Objetivo: '{fecha_objetivo}'")
                        
                        # Extracción de datos (usando tu función existente)
                        expediente, causa, juzgado, codigo = extraer_datos_mejorado(driver, fila, i+1, aria_label)
                        
                        # Validar y agregar si los datos son válidos
                        if expediente != "No disponible":
                            exp_upper = expediente.upper().strip()
                            if exp_upper not in expedientes_procesados:
                                fila_datos = {
                                    'expediente': expediente,
                                    'juzgado': juzgado,
                                    'causa': causa,
                                    'estado': 'Notificación',
                                    'fecha': fecha_extraida,
                                    'fuente': 'Notificaciones',
                                    'codigo': codigo
                                }
                                filas_notificaciones.append(fila_datos)
                                expedientes_procesados.add(exp_upper)
                                filas_pagina_actual += 1

                                print(f"   🎉 NUEVA NOTIFICACIÓN AGREGADA:")
                                print(f"       Expediente: {expediente}")
                                print(f"       Causa: {causa[:50]}...")
                            else:
                                print(f"   ⚠️ EXPEDIENTE YA PROCESADO: {expediente}")
                        
                except Exception as e:
                    print(f"   ❌ Error procesando fila {i+1}: {e}")
                    continue
            
            print(f"✅ Página {pagina_actual}: {filas_pagina_actual} nuevas notificaciones encontradas")
            print(f"📊 Total acumulado: {len(filas_notificaciones)} notificaciones")
            
            # Intentar ir a la siguiente página
            if ir_siguiente_pagina_notificaciones(driver):
                pagina_actual += 1
                time.sleep(2)  # Pequeña pausa inicial para que empiece el reload
                _esperar_filas_estables(driver)
            else:
                print("🛑 No hay más páginas o no se pudo navegar. Finalizando paginación.")
                break
                
        except Exception as e:
            print(f"❌ Error en página {pagina_actual}: {e}")
            break
    
    return filas_notificaciones

def paginar_tabla_expediente_mejorado(driver, tabla_id, textos_extraidos):
    """
    FUNCIÓN CORREGIDA: Función mejorada para paginar tablas de expedientes (basada en tu código viejo que funcionaba)
    """
    print(f"📄 Iniciando paginación para tabla: {tabla_id}")
    
    # Extraer primera página
    extraer_info_pagina(driver, tabla_id, textos_extraidos)
    
    pagina_actual = 1
    tabla_referencia = None
    
    try:
        tabla_referencia = driver.find_element(By.ID, tabla_id)
    except:
        print(f"❌ No se pudo obtener referencia de tabla {tabla_id}")
        return
    
    while tabla_referencia:
        try:
            # Buscar paginador según el tipo de tabla
            if "action-table" in tabla_id:
                # Para actuaciones recientes
                paginador = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "ul.pagination"))
                )
                siguiente_li = paginador.find_element(By.CSS_SELECTOR, "li.active + li:not(.disabled)")
                siguiente_link = siguiente_li.find_element(By.TAG_NAME, "a")
                
            elif "historic-table" in tabla_id:
                # Para actuaciones históricas
                paginador = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.ui-paginator-bottom"))
                )
                siguiente_link = paginador.find_element(By.CSS_SELECTOR, "span.ui-paginator-next:not(.ui-state-disabled)")
            
            else:
                print(f"⚠️ Tipo de tabla no reconocido: {tabla_id}")
                break
            
            pagina_actual += 1
            print(f"🔄 Navegando a página {pagina_actual} de {tabla_id}...")
            
            # Hacer clic en siguiente
            driver.execute_script("arguments[0].click();", siguiente_link)
            
            # Esperar a que la tabla anterior se vuelva obsoleta
            WebDriverWait(driver, 15).until(EC.staleness_of(tabla_referencia))
            
            # Esperar a que aparezca la nueva tabla
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.ID, tabla_id)))
            
            # Tiempo adicional para carga completa
            if "historic-table" in tabla_id:
                time.sleep(3)  # Más tiempo para históricas
            else:
                time.sleep(2)
            
            # Extraer datos de la nueva página
            extraer_info_pagina(driver, tabla_id, textos_extraidos)
            
            # Actualizar referencia de tabla
            tabla_referencia = driver.find_element(By.ID, tabla_id)
            
        except Exception as e:
            print(f"🛑 Fin de paginación para {tabla_id}: {e}")
            break
    
    print(f"✅ Paginación completada para {tabla_id}")

# ==================== FUNCIONES PRINCIPALES CORREGIDAS ====================

def filtrar_por_fecha(fecha_objetivo, paginas_a_procesar, usuario, password, headless=True, filas_deox=10, gemini_api_key=None, captcha_api_key=None):
    """Función principal de extracción de expedientes (VERSIÓN CORREGIDA)"""
    
    options = Options()
    if headless:
        print("🔇 Modo HEADLESS activado - El navegador no será visible")
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-images")
    else:
        print("🖥️ Modo NORMAL activado - El navegador será visible")
        options.add_argument("--start-maximized")
    
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    if headless:
        driver.set_page_load_timeout(60)
        driver.implicitly_wait(10)
    
    try:
        # Login y navegación inicial - Navegar al portal para obtener state/nonce frescos
        url_portal = "https://portalpjn.pjn.gov.ar/"
        
        print(f"🌐 Navegando al portal PJN para obtener redirección de login...")
        driver.get(url_portal)
        
        # Esperar a que redirija al SSO automáticamente (el portal genera state/nonce frescos)
        print(f"🔄 Esperando redirección al SSO...")
        time.sleep(3)
        
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "username")))
        
        print(f"👤 Usando usuario: {usuario}")
        print(f"🔐 Realizando login...")
        driver.find_element(By.ID, "username").send_keys(usuario)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.ID, "kc-login").click()
        
        print("⏳ Esperando confirmación de login...")
        time.sleep(5)
        
        print(f"📅 INFORMACIÓN SOBRE FECHAS:")
        print(f"   Fecha ingresada: {fecha_objetivo}")
        print(f"   Fecha normalizada para comparación: {normalizar_fecha_para_comparacion(fecha_objetivo)}")
        print(f"   Fecha para Excel: {normalizar_fecha_para_excel(fecha_objetivo)}")
        print("=" * 80)
        
        # Procesar Consultas
        print("🔍 Procesando consultas...")
        filas_consultas = procesar_consultas(driver, fecha_objetivo, paginas_a_procesar)
        
        # CORREGIDO: Procesar Notificaciones con configuración de 30 elementos
        print("\n🔔 Procesando notificaciones con configuración mejorada...")
        filas_notificaciones = procesar_notificaciones_mejorado(driver, fecha_objetivo, filas_consultas)
        
        # Procesar DEOX
        print(f"\n📋 Procesando DEOX (máximo {filas_deox} filas)...")
        filas_deox_resultado = procesar_deox(driver, fecha_objetivo, filas_consultas + filas_notificaciones, filas_deox)
        
        # Combinar todos los datos
        todos_los_datos = filas_consultas + filas_notificaciones + filas_deox_resultado
        
        # Normalizar fechas para Excel
        for fila in todos_los_datos:
            fila['fecha'] = normalizar_fecha_para_excel(fila['fecha'])
        
        # Extraer números y años de expedientes
        print(f"\n🔢 Extrayendo números y años de expedientes...")
        for i, fila in enumerate(todos_los_datos):
            numero, ano = extraer_numero_y_ano_expediente(fila['expediente'])
            fila['numero_expediente'] = numero
            fila['ano_expediente'] = ano
            
            if numero and ano:
                print(f"   ✅ Fila {i+1}: {fila['expediente']} → Número: {numero}, Año: {ano}")
            else:
                print(f"   ⚠️ Fila {i+1}: No se pudo extraer de {fila['expediente']}")
        
        # Guardar todos los datos
        guardar_todos_en_excel(todos_los_datos, normalizar_fecha_para_excel(fecha_objetivo))
        
        # Mostrar resumen final
        mostrar_resumen_final(filas_consultas, filas_notificaciones, filas_deox_resultado, fecha_objetivo)
        
    except Exception as e:
        print(f"❌ Error durante la ejecución: {str(e)}")
        if headless:
            print("💡 Sugerencia: Si hay problemas en modo headless, pruebe con modo normal")
    finally:
        print("🔄 Cerrando navegador...")
        driver.quit()

def procesar_notificaciones_mejorado(driver, fecha_objetivo, filas_consultas):
    """
    FUNCIÓN CORREGIDA: Procesa la sección de notificaciones con configuración de 30 elementos
    """
    ventana_original = driver.current_window_handle
    pestaña_abierta = False
    filas_notificaciones = []
    try:
        # Buscar y hacer clic en el elemento de Notificaciones
        print("🔍 Buscando sección Notificaciones...")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, "//span[contains(@class, 'MuiTypography-root') and text()='Notificaciones']")
        ))
        driver.find_element(By.XPATH, "//span[contains(@class, 'MuiTypography-root') and text()='Notificaciones']").click()

        # Esperar la apertura y cambiar a la nueva ventana
        print("⏳ Esperando apertura de nueva pestaña...")
        WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > 1)
        driver.switch_to.window(driver.window_handles[-1])
        pestaña_abierta = True

        print("🔔 Navegando a la sección de notificaciones...")

        # Esperar a que la tabla se cargue y estabilice
        print("⏳ Esperando carga de tabla de notificaciones...")
        _esperar_filas_estables(driver)

        # Configurar 30 filas por página
        print("⚙️ Configurando 30 filas por página...")
        try:
            select_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@aria-label='filas por página']"))
            )
            WebDriverWait(driver, 10).until(
                lambda d: not d.find_element(By.XPATH, "//select[@aria-label='filas por página']").get_attribute('disabled')
            )
            sel = Select(select_element)
            valor_actual = sel.first_selected_option.get_attribute('value')
            if valor_actual != '30':
                try:
                    sel.select_by_value("30")
                except Exception:
                    opciones = [o for o in sel.options if not o.get_attribute('disabled')]
                    if opciones:
                        sel.select_by_value(opciones[-1].get_attribute('value'))
                print("✅ Filas por página configuradas — esperando recarga...")
                _esperar_filas_estables(driver)
            else:
                print("✅ Ya está en 30 filas por página")
        except Exception as e:
            print(f"⚠️ No se pudo configurar filas por página: {e}")

        # Procesar las filas de notificaciones con paginación mejorada
        filas_notificaciones = procesar_filas_notificaciones_con_paginacion(driver, fecha_objetivo, filas_consultas)

        print(f"✅ Procesamiento de notificaciones completado. Encontradas {len(filas_notificaciones)} notificaciones nuevas.")

    except Exception as e:
        print(f"❌ Error procesando notificaciones: {str(e)}")
    finally:
        # Siempre cerrar la pestaña de notificaciones y volver a la ventana original
        try:
            if pestaña_abierta and len(driver.window_handles) > 1:
                driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(ventana_original)
            print("🔄 Volviendo a la ventana principal...")
        except Exception:
            pass

    return filas_notificaciones

def analizar_expedientes_individuales(usuario, password, headless=True, gemini_api_key=None, captcha_api_key=None):
    """
    FUNCIÓN PRINCIPAL MODIFICADA: Análisis individual de expedientes extrayendo info del HTML
    """
    
    # Configurar Tesseract (mantenido por compatibilidad)
    configurar_tesseract()
    
    # Configurar Gemini
    if gemini_api_key:
        try:
            genai.configure(api_key=gemini_api_key)
            print("✅ API de Gemini configurada correctamente.")
        except Exception as e:
            print(f"❌ Error configurando la API de Gemini: {e}")
            return
    else:
        print("❌ API Key de Gemini no proporcionada.")
        return
    
    # Archivos
    archivo_entrada = 'expedientes.xlsx'
    archivo_salida = 'expedientes_analizados.xlsx'
    
    # Cargar expedientes
    try:
        df_input = pd.read_excel(archivo_entrada)
        print(f"📊 Archivo '{archivo_entrada}' cargado con {len(df_input)} filas.")
    except FileNotFoundError:
        print(f"❌ Error: El archivo '{archivo_entrada}' no se encontró.")
        return
    except Exception as e:
        print(f"❌ Error leyendo el archivo '{archivo_entrada}': {e}")
        return
    
    # Definir columnas de salida
    columnas_salida = list(df_input.columns) + ['Resumen del Caso', 'Estado Análisis']
    
    # Cargar resultados previos
    try:
        df_resultado = pd.read_excel(archivo_salida, dtype={'Número': str, 'Año': str})
        df_resultado['Número'] = df_resultado['Número'].astype(str).str.replace(r'\.0$', '', regex=True)
        df_resultado['Año'] = df_resultado['Año'].astype(str).str.replace(r'\.0$', '', regex=True)
        print(f"📊 Archivo de resultados previos '{archivo_salida}' cargado con {len(df_resultado)} filas.")
        
        # Asegurar columnas
        for col in columnas_salida:
            if col not in df_resultado.columns:
                print(f"➕ Añadiendo columna faltante: {col}")
                df_resultado[col] = ''
                
    except FileNotFoundError:
        print(f"📄 Archivo '{archivo_salida}' no encontrado. Se creará uno nuevo.")
        df_resultado = pd.DataFrame(columns=columnas_salida)
    except Exception as e:
        print(f"⚠️ Error cargando resultados previos: {e}. Se iniciará con DataFrame vacío.")
        df_resultado = pd.DataFrame(columns=columnas_salida)
    
    # Configurar WebDriver
    options = Options()
    if headless:
        options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        print("✅ WebDriver iniciado correctamente.")
    except Exception as e:
        print(f"❌ Error iniciando WebDriver: {e}")
        return
    
    # Procesar cada expediente
    expedientes_procesados = 0
    
    for index, row in df_input.iterrows():
        
        # Obtener número y año
        numero = str(row.get('Número', '')).split('.')[0]
        ano = str(row.get('Año', '')).split('.')[0]
        
        if not numero or not ano:
            print(f"⚠️ Error en Fila {index+1}: Falta número o año. Saltando.")
            continue
        
        expediente_id = f"{numero}/{ano}"
        print(f"\n--- 🔍 Procesando Expediente: {expediente_id} (Fila {index+1}/{len(df_input)}) ---")
        
        # Verificar si ya fue procesado
        expedientes_existentes = (df_resultado['Número'].astype(str) + '/' + df_resultado['Año'].astype(str)).tolist()
        if expediente_id in expedientes_existentes:
            print(f"✅ Expediente {expediente_id} ya procesado. Saltando.")
            continue
        
        # Variables para este expediente
        resumen_caso = ""
        estado_analisis = "Pendiente"
        
        try:
            # Abrir página de consulta
            driver.get("https://scw.pjn.gov.ar/scw/home.seam")
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "formPublica:numero")))
            print("📄 Página de consulta cargada.")
            time.sleep(1)
            
            # Ingresar datos
            driver.find_element(By.ID, "formPublica:numero").send_keys(numero)
            driver.find_element(By.ID, "formPublica:anio").send_keys(ano)
            Select(driver.find_element(By.ID, "formPublica:camaraNumAni")).select_by_value("24")
            print(f"📝 Datos del expediente {expediente_id} ingresados.")
            time.sleep(1)
            
            # Resolver CAPTCHA
            sitekey = "6LcTJ1kUAAAAAJT1Xqu3gzANPfCbQG0nke9O5b6K"
            captcha_solution = solve_recaptcha(captcha_api_key, sitekey, driver.current_url)
            
            if not captcha_solution:
                print("❌ Error: Falló la resolución del CAPTCHA.")
                estado_analisis = "Error: CAPTCHA falló"
                resumen_caso = "No se pudo resolver el CAPTCHA para acceder al expediente."
            else:
                # Inyectar solución y buscar
                print("🔓 Inyectando solución del CAPTCHA...")
                driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = arguments[0];", captcha_solution)
                time.sleep(1)
                
                print("🔍 Haciendo clic en buscar...")
                buscar_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "formPublica:buscarPorNumeroButton")))
                buscar_button.click()
                print("⏳ Esperando resultados...")
                time.sleep(5)
                
                # Verificar si el expediente existe
                try:
                    messages_element = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.ui-messages-error-summary, div.ui-messages-info-summary")))
                    msg = messages_element.text
                    print(f"📋 Mensaje encontrado: {msg}")
                    
                    if "Expediente inexistente" in msg or "no se registran actuaciones" in msg:
                        print(f"❌ Expediente {expediente_id} no encontrado.")
                        estado_analisis = "Expediente Inexistente"
                        resumen_caso = "El expediente no existe o no tiene actuaciones registradas."
                    else:
                        estado_analisis = "Error: Mensaje desconocido"
                        resumen_caso = f"Mensaje del sistema: {msg}"
                        
                except Exception:
                    print("✅ No se encontraron mensajes de error. Procediendo a extraer datos HTML.")
                    
                    # ==================== NUEVA FUNCIONALIDAD: Extraer información del HTML ====================
                    print("📄 Extrayendo información básica del expediente...")
                    datos_basicos = extraer_datos_basicos_expediente(driver)
                    
                    print("👥 Extrayendo información de intervinientes...")
                    datos_intervinientes = hacer_click_intervinientes_y_extraer(driver)
                    
                    # Combinar todos los datos
                    datos_completos = {**datos_basicos, **datos_intervinientes}
                    
                    print(f"📊 Datos extraídos del HTML:")
                    for key, value in datos_completos.items():
                        print(f"   {key}: {value}")
                    
                    # Analizar con IA usando la información extraída del HTML
                    print("🤖 Enviando información extraída a Gemini para análisis JSON...")
                    respuesta_ia = analyze_legal_documents_html(datos_completos, gemini_api_key)
                    
                    if respuesta_ia and respuesta_ia.strip():
                        # Procesar respuesta JSON
                        try:
                            # Limpiar la respuesta por si tiene texto adicional
                            respuesta_limpia = respuesta_ia.strip()
                            
                            # Buscar el JSON en la respuesta
                            inicio_json = respuesta_limpia.find('{')
                            fin_json = respuesta_limpia.rfind('}') + 1
                            
                            if inicio_json != -1 and fin_json != -1:
                                json_str = respuesta_limpia[inicio_json:fin_json]
                                datos_json = json.loads(json_str)
                                
                                print(f"📋 JSON recibido de la IA:")
                                for key, value in datos_json.items():
                                    print(f"   {key}: {value}")
                                
                                # Crear registro con columnas separadas
                                registro = row.to_dict()
                                
                                # Agregar datos del análisis en columnas separadas
                                registro['Expediente_Numero'] = datos_json.get('expediente_numero', 'No disponible')
                                registro['Jurisdiccion_Normalizada'] = datos_json.get('jurisdiccion_normalizada', 'No disponible')
                                registro['Juzgado_Normalizado'] = datos_json.get('juzgado', 'No disponible')
                                registro['Secretaria'] = datos_json.get('secretaria', 'No disponible')
                                registro['Situacion_Actual'] = datos_json.get('situacion_actual', 'No disponible')
                                registro['Tipo_Proceso'] = datos_json.get('tipo_proceso', 'No disponible')
                                registro['Actor_Principal'] = datos_json.get('actor_principal', 'No disponible')
                                registro['Demandado_Principal'] = datos_json.get('demandado_principal', 'No disponible')
                                registro['Objeto_Demanda'] = datos_json.get('objeto_demanda', 'No disponible')
                                registro['Letrado_Actor'] = datos_json.get('letrado_actor', 'No disponible')
                                registro['Matricula_Letrado'] = datos_json.get('matricula_letrado', 'No disponible')
                                registro['CUIT_CUIL_Letrado'] = datos_json.get('cuit_cuil_letrado', 'No disponible')
                                registro['Observaciones'] = datos_json.get('observaciones', 'No disponible')
                                registro['Estado_Analisis'] = "Completado - JSON"
                                
                                print("✅ Análisis completado exitosamente con JSON estructurado.")
                                print(f"   📊 Datos extraídos:")
                                print(f"      - Actor: {registro['Actor_Principal']}")
                                print(f"      - Letrado: {registro['Letrado_Actor']}")
                                print(f"      - Tipo Proceso: {registro['Tipo_Proceso']}")
                                
                            else:
                                # Si no es JSON válido, guardar como error
                                registro = row.to_dict()
                                registro['Estado_Analisis'] = "Error: JSON inválido"
                                registro['Observaciones'] = "La IA no devolvió un JSON válido"
                                
                        except json.JSONDecodeError as e:
                            print(f"❌ Error decodificando JSON: {e}")
                            registro = row.to_dict()
                            registro['Estado_Analisis'] = "Error: JSON inválido"
                            registro['Observaciones'] = f"Error JSON: {str(e)}"
                            
                    else:
                        registro = row.to_dict()
                        registro['Estado_Analisis'] = "Error: Análisis falló"
                        registro['Observaciones'] = "El análisis con IA no pudo completarse"
            
            # Agregar al DataFrame
            df_resultado = pd.concat([df_resultado, pd.DataFrame([registro])], ignore_index=True)
            expedientes_procesados += 1
        
            # Guardar progreso cada 3 expedientes
            if expedientes_procesados % 3 == 0:
                try:
                    df_resultado_guardar = df_resultado.reindex(columns=columnas_salida, fill_value='')
                    df_resultado_guardar.to_excel(archivo_salida, index=False)
                    print(f"💾 Progreso guardado ({expedientes_procesados} procesados).")
                except Exception as e_save:
                    print(f"⚠️ Error al guardar progreso: {e_save}")
            
        except Exception as e:
            print(f"❌ Error crítico procesando expediente {expediente_id}: {e}")
            traceback.print_exc()
            
            # Registrar error
            registro_error = row.to_dict()
            registro_error['Resumen del Caso'] = f"Error crítico: {str(e)}"
            registro_error['Estado Análisis'] = "Error Crítico"
            df_resultado = pd.concat([df_resultado, pd.DataFrame([registro_error])], ignore_index=True)
            expedientes_procesados += 1
        
        finally:
            print("⏳ Esperando 3 segundos antes del siguiente expediente...")
            time.sleep(3)
    
    # Finalización
    if driver:
        driver.quit()
    
    # Guardar resultados finales usando la nueva función
    try:
        guardar_datos_en_excel_analizados(df_resultado, archivo_salida)
        print(f"💾 Guardado final de '{archivo_salida}' realizado con columnas separadas.")
    except Exception as e_final:
        print(f"❌ Error en el guardado final: {e_final}")
    
    print("\n" + "=" * 60)
    print("✅ ANÁLISIS COMPLETADO")
    print(f"📊 Total expedientes procesados: {expedientes_procesados}")
    print(f"📂 Resultados guardados en: {archivo_salida}")
    print("💡 NUEVA VERSIÓN: Extrae datos HTML con IA")
    print("=" * 60)

# ==================== FUNCIONES AUXILIARES (CÓDIGO ORIGINAL) ====================

def procesar_consultas(driver, fecha_objetivo, filas_consultas):
    """Procesa la sección de consultas con paginación mejorada"""
    print("🔍 Buscando sección Consultas...")
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Consultas']")))
    driver.find_element(By.XPATH, "//span[text()='Consultas']").click()
    
    # Esperar la apertura y cambiar a la nueva ventana
    print("⏳ Esperando apertura de nueva pestaña...")
    WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > 1)
    driver.switch_to.window(driver.window_handles[-1])
    
    # Configurar ordenamiento por fecha
    print("⚙️ Configurando ordenamiento por fecha...")
    WebDriverWait(driver, 15).until(EC.presence_of_element_located(
        (By.ID, "j_idt150:order_by_form:camara")
    ))
    
    select_element = Select(driver.find_element(By.ID, "j_idt150:order_by_form:camara"))
    select_element.select_by_value("FECHA")
    
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.XPATH, "//a[contains(text(), 'Ordenar')]")
    )).click()
    
    time.sleep(5)  # Esperar más tiempo para que se ordene
    
    # Variables para controlar la paginación
    filas_encontradas = []
    
    print(f"🔍 Buscando expedientes con fecha: {fecha_objetivo}")
    print(f"📋 Procesando las primeras {filas_consultas} páginas")
    print("=" * 80)
    
    for pagina_actual in range(1, filas_consultas + 1):
        print(f"\n📄 === PROCESANDO PÁGINA {pagina_actual} DE {filas_consultas} ===")
        
        # Verificar que estamos en la página correcta
        try:
            # Buscar indicador de página actual
            indicadores_pagina = driver.find_elements(By.XPATH, "//span[contains(@class, 'current-page') or contains(@class, 'active-page')]")
            
            if indicadores_pagina:
                pagina_mostrada = indicadores_pagina[0].text.strip()
                print(f"📍 Página mostrada en interfaz: {pagina_mostrada}")
        except:
            print(f"📍 No se pudo determinar número de página actual")
        
        # Contar filas antes de procesar
        try:
            filas_tabla = driver.find_elements(By.XPATH, "//table//tbody//tr")
            print(f"📊 Filas encontradas en la tabla: {len(filas_tabla)}")
        except:
            print(f"⚠️ No se pudieron contar las filas de la tabla")
        
        # Procesar filas de la página actual
        filas_pagina = procesar_filas_pagina(driver, fecha_objetivo)
        filas_encontradas.extend(filas_pagina)
        
        print(f"✅ Página {pagina_actual}: {len(filas_pagina)} filas con fecha {fecha_objetivo}")
        print(f"📊 Total acumulado: {len(filas_encontradas)} filas")
        
        # Si no es la última página, intentar ir a la siguiente
        if pagina_actual < filas_consultas:
            print(f"\n🔄 Intentando navegar a página {pagina_actual + 1}...")
            
            if ir_siguiente_pagina(driver):
                print(f"✅ Navegación exitosa a página {pagina_actual + 1}")
                time.sleep(5)  # Esperar más tiempo para que cargue
            else:
                print(f"❌ No se pudo navegar a la página {pagina_actual + 1}")
                print(f"🛑 Deteniendo procesamiento en página {pagina_actual}")
                break
        else:
            print(f"ℹ️ Última página procesada: {pagina_actual}")
    
    print(f"\n✅ Procesamiento de consultas completado.")
    print(f"📊 Total de filas encontradas: {len(filas_encontradas)}")
    
    # Cerrar pestaña de consultas y volver a la original
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    print("🔄 Cerrada pestaña de consultas, volviendo a la original...")
    
    return filas_encontradas

def procesar_deox(driver, fecha_objetivo, filas_procesadas, max_filas=10):
    """Procesa la sección de DEOX - ACTUALIZADO para nueva estructura MUI"""
    ventana_original = driver.current_window_handle
    pestaña_abierta = False
    filas_deox = []
    try:
        # Buscar y hacer clic en el elemento de DEOX
        print("🔍 Buscando sección DEOX...")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, "//span[contains(@class, 'MuiTypography-root') and text()='DEOX']")
        ))
        driver.find_element(By.XPATH, "//span[contains(@class, 'MuiTypography-root') and text()='DEOX']").click()

        # Esperar la apertura y cambiar a la nueva ventana
        print("⏳ Esperando apertura de nueva pestaña...")
        WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > 1)
        driver.switch_to.window(driver.window_handles[-1])
        pestaña_abierta = True

        print("📋 Navegando a la sección de DEOX...")

        # Esperar a que la tabla se cargue (nueva estructura MUI)
        print("⏳ Esperando carga de tabla DEOX...")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, "//tr[contains(@class, 'MuiBox-root') and @role='row']")
        ))
        time.sleep(3)

        # Configurar 30 filas por página antes de extraer
        print("⚙️ Configurando 30 filas por página en DEOX...")
        try:
            select_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@aria-label='filas por página']"))
            )
            WebDriverWait(driver, 10).until(
                lambda d: not d.find_element(By.XPATH, "//select[@aria-label='filas por página']").get_attribute('disabled')
            )
            sel = Select(select_element)
            try:
                sel.select_by_value("30")
                print("✅ Seleccionado 30 filas por página")
            except Exception:
                opciones_habilitadas = [o for o in sel.options if not o.get_attribute('disabled')]
                if opciones_habilitadas:
                    sel.select_by_value(opciones_habilitadas[-1].get_attribute('value'))
                    print(f"✅ Seleccionado {opciones_habilitadas[-1].get_attribute('value')} filas por página")
            print("⏳ Esperando 5 segundos para que recargue la tabla...")
            time.sleep(5)
        except Exception as e:
            print(f"⚠️ No se pudo configurar 30 filas por página: {e}")
            print("   Continuando con la cantidad por defecto...")

        # Procesar las filas de DEOX con límite
        filas_deox = procesar_filas_deox(driver, fecha_objetivo, filas_procesadas, max_filas)

        print(f"✅ Procesamiento de DEOX completado. Encontradas {len(filas_deox)} registros nuevos.")

    except Exception as e:
        print(f"❌ Error procesando DEOX: {str(e)}")
    finally:
        # Siempre cerrar la pestaña de DEOX y volver a la ventana original
        try:
            if pestaña_abierta and len(driver.window_handles) > 1:
                driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(ventana_original)
            print("🔄 Volviendo a la ventana principal...")
        except Exception:
            pass

    return filas_deox

def procesar_filas_pagina(driver, fecha_objetivo):
    """Procesa todas las filas de la página actual y filtra por fecha"""
    filas_encontradas = []
    
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//table//tbody//tr"))
        )
        
        filas = driver.find_elements(By.XPATH, "//table//tbody//tr")
        
        for i, fila in enumerate(filas):
            try:
                celdas = fila.find_elements(By.TAG_NAME, "td")
                
                if len(celdas) >= 5:
                    expediente = celdas[0].text.strip()
                    juzgado = celdas[1].text.strip()
                    causa = celdas[2].text.strip()
                    estado = celdas[3].text.strip()
                    fecha = celdas[4].text.strip()
                    
                    # Usar la función de comparación mejorada
                    if comparar_fechas_mejorado(fecha, fecha_objetivo):
                        fila_datos = {
                            'expediente': expediente,
                            'juzgado': juzgado,
                            'causa': causa,
                            'estado': estado,
                            'fecha': fecha,
                            'fuente': 'Consultas'
                        }
                        filas_encontradas.append(fila_datos)
                        
                        print(f"   ✓ COINCIDENCIA ENCONTRADA:")
                        print(f"     Fecha en tabla: '{fecha}' vs Fecha objetivo: '{fecha_objetivo}'")
                        print(f"     {expediente} | {juzgado} | {causa[:50]}... | {estado}")
                        
            except Exception as e:
                print(f"   ⚠️ Error procesando fila {i+1}: {str(e)}")
                continue
                
    except Exception as e:
        print(f"   ❌ Error procesando página: {str(e)}")
    
    return filas_encontradas

def procesar_filas_deox(driver, fecha_objetivo, filas_procesadas, max_filas=10):
    """Procesa las filas de DEOX - ACTUALIZADO para nueva estructura MUI"""
    filas_deox = []
    
    try:
        filas = driver.find_elements(By.XPATH, "//tr[contains(@class, 'MuiBox-root') and @role='row']")
        total_filas_disponibles = len(filas)
        filas_a_procesar = min(max_filas, total_filas_disponibles)
        
        print(f"📋 Encontradas {total_filas_disponibles} filas de DEOX disponibles")
        print(f"🎯 Procesando las primeras {filas_a_procesar} filas (límite: {max_filas})")
        
        expedientes_procesados = set()
        for fila_proc in filas_procesadas:
            if 'expediente' in fila_proc:
                expedientes_procesados.add(fila_proc['expediente'].strip().upper())
        
        # Procesar solo las primeras max_filas filas
        for i in range(filas_a_procesar):
            try:
                fila = filas[i]
                
                # Extraer fecha del aria-label (formato: "Fecha Recibido: DD/M/YYYY")
                aria_label = fila.get_attribute("aria-label")
                fecha_extraida = None
                
                if aria_label:
                    # Patrón para "Fecha Recibido: DD/M/YYYY" o "Fecha Recibido: D/M/YYYY"
                    match_fecha = re.search(r'Fecha Recibido:\s*(\d{1,2}/\d{1,2}/\d{4})', aria_label)
                    if match_fecha:
                        fecha_extraida = match_fecha.group(1)
                
                print(f"   🔍 Fila DEOX {i+1}/{filas_a_procesar}:")
                
                if not fecha_extraida:
                    print(f"       ⚠️ No se pudo extraer fecha del aria-label")
                    continue
                
                print(f"       📅 Fecha extraída: {fecha_extraida}")
                
                # Usar la función de comparación mejorada
                if comparar_fechas_mejorado(fecha_extraida, fecha_objetivo):
                    print(f"       ✅ COINCIDENCIA DE FECHA DEOX:")
                    print(f"           Extraída: '{fecha_extraida}' vs Objetivo: '{fecha_objetivo}'")
                    
                    # Extraer datos del TD con descripcionMobile
                    expediente = "No disponible"
                    causa = "No disponible"
                    juzgado = "No disponible"
                    id_registro = "No disponible"
                    
                    try:
                        # Estrategia 1: Buscar TD por data-key conteniendo 'descripcion'
                        td_datos = None
                        selectores_td = [
                            "td[data-key*='descripcionMobile']",
                            "td[data-key*='descripcion']",
                            "td[data-key*='Descripcion']",
                        ]
                        for selector in selectores_td:
                            try:
                                td_datos = fila.find_element(By.CSS_SELECTOR, selector)
                                break
                            except:
                                continue
                        
                        # Estrategia 2: Si no encontramos por data-key, buscar el TD con role='gridcell' que tenga más contenido
                        if not td_datos:
                            tds_gridcell = fila.find_elements(By.CSS_SELECTOR, "td[role='gridcell']")
                            if tds_gridcell:
                                # Elegir el TD con más texto (el de descripción es el más largo)
                                td_datos = max(tds_gridcell, key=lambda td: len(td.text))
                        
                        # Estrategia 3: Buscar cualquier TD que no sea el de fecha ni el de etiqueta
                        if not td_datos:
                            all_tds = fila.find_elements(By.TAG_NAME, "td")
                            for td in all_tds:
                                texto_td = td.text.strip()
                                # El TD de descripción tiene el texto más largo
                                if len(texto_td) > 20:
                                    td_datos = td
                                    break
                        
                        if td_datos:
                            # Extraer ID/número de oficio (primer <p>)
                            try:
                                p_elem = td_datos.find_element(By.TAG_NAME, "p")
                                id_registro = p_elem.text.strip()
                            except:
                                pass
                            
                            # Extraer expediente y causa del primer div.ellipsed-cell-div
                            divs_ellipsed = td_datos.find_elements(By.CSS_SELECTOR, "div.ellipsed-cell-div")
                            
                            if len(divs_ellipsed) >= 1:
                                # Primer div: expediente + causa
                                spans_primer_div = divs_ellipsed[0].find_elements(By.TAG_NAME, "span")
                                for span in spans_primer_div:
                                    texto_span = span.text.strip()
                                    if not texto_span:
                                        continue
                                    # El expediente tiene formato "FSA NNNNN/YYYY" - solo capturar el match
                                    exp_match = re.match(r'^[A-Z]{2,5}\s*\d{1,9}/\d{4}', texto_span)
                                    if exp_match:
                                        expediente = exp_match.group(0)
                                    # La causa contiene "c/" o "s/" y es más larga
                                    elif ('c/' in texto_span.lower() or 's/' in texto_span.lower()) and len(texto_span) > 10:
                                        causa = texto_span
                            
                            if len(divs_ellipsed) >= 2:
                                # Segundo div: juzgado/origen
                                spans_segundo_div = divs_ellipsed[1].find_elements(By.TAG_NAME, "span")
                                for span in spans_segundo_div:
                                    texto_span = span.text.strip()
                                    if texto_span and len(texto_span) > 5:
                                        juzgado = texto_span
                                        break
                                
                                # Si no encontramos en spans, buscar en divs internos
                                if juzgado == "No disponible":
                                    divs_internos = divs_ellipsed[1].find_elements(By.TAG_NAME, "div")
                                    for div_int in divs_internos:
                                        texto_div = div_int.text.strip()
                                        if texto_div and len(texto_div) > 5:
                                            juzgado = texto_div
                                            break
                            
                            # Si no encontramos divs ellipsed, buscar directamente en spans del TD
                            if expediente == "No disponible":
                                all_spans = td_datos.find_elements(By.TAG_NAME, "span")
                                for span in all_spans:
                                    texto_span = span.text.strip()
                                    if not texto_span:
                                        continue
                                    exp_match_fb = re.match(r'^[A-Z]{2,5}\s*\d{1,9}/\d{4}', texto_span) if expediente == "No disponible" else None
                                    if exp_match_fb:
                                        expediente = exp_match_fb.group(0)
                                    elif causa == "No disponible" and ('c/' in texto_span.lower() or 's/' in texto_span.lower()) and len(texto_span) > 10:
                                        causa = texto_span
                                    elif juzgado == "No disponible" and ('JUZGADO' in texto_span.upper() or 'CAMARA' in texto_span.upper() or 'TRIBUNAL' in texto_span.upper() or 'SECRETARIA' in texto_span.upper()):
                                        juzgado = texto_span
                        else:
                            print(f"       ⚠️ No se encontró TD de datos en la fila")
                        
                    except Exception as e:
                        print(f"       ⚠️ Error extrayendo datos del TD: {e}")
                    
                    # Fallback: extraer del aria-label si no se pudo del TD
                    if aria_label:
                        # Extraer ID de oficio del aria-label
                        if id_registro == "No disponible":
                            oficio_match = re.search(r'Oficio Número:\s*(\d+)', aria_label)
                            if oficio_match:
                                id_registro = oficio_match.group(1).strip()
                        
                        # Extraer juzgado/origen del aria-label
                        if juzgado == "No disponible":
                            origen_match = re.search(r'Origen:\s*(.+?)\s*-\s*Fecha', aria_label)
                            if origen_match:
                                juzgado = origen_match.group(1).strip()
                        
                        # El campo "Expediente:" del aria-label contiene la carátula/causa, no el número
                        if causa == "No disponible":
                            exp_caratula_match = re.search(r'Expediente:\s*(.+?)$', aria_label)
                            if exp_caratula_match:
                                causa = exp_caratula_match.group(1).strip()
                        
                        # Buscar número de expediente (FSA XXXX/YYYY) en todo el texto de la fila
                        if expediente == "No disponible":
                            try:
                                texto_fila_completo = fila.text
                                exp_num_match = re.search(r'([A-Z]{2,5}\s*\d{1,9}/\d{4})', texto_fila_completo)
                                if exp_num_match:
                                    expediente = exp_num_match.group(1)
                                    print(f"       ✅ Expediente extraído del texto de la fila: {expediente}")
                            except:
                                pass
                    
                    if expediente != "No disponible":
                        expediente_upper = expediente.upper().strip()
                        
                        if expediente_upper not in expedientes_procesados:
                            fila_datos = {
                                'expediente': expediente,
                                'juzgado': juzgado,
                                'causa': causa,
                                'estado': 'DEOX',
                                'fecha': fecha_extraida,
                                'fuente': 'DEOX',
                                'codigo': id_registro
                            }
                            filas_deox.append(fila_datos)
                            expedientes_procesados.add(expediente_upper)
                            
                            print(f"       🎉 NUEVO REGISTRO DEOX AGREGADO:")
                            print(f"           Expediente: {expediente}")
                            print(f"           Causa: {causa}")
                            print(f"           Juzgado: {juzgado}")
                        else:
                            print(f"       ⚠️ EXPEDIENTE YA PROCESADO: {expediente}")
                    else:
                        print(f"       ❌ No se pudo extraer expediente válido")
                else:
                    print(f"       ⏭️ Fecha no coincide: '{fecha_extraida}' vs '{fecha_objetivo}'")
                        
            except Exception as e:
                print(f"   ❌ Error procesando fila DEOX {i+1}: {e}")
                continue
        
        if total_filas_disponibles > max_filas:
            print(f"ℹ️ Se procesaron {filas_a_procesar} de {total_filas_disponibles} filas disponibles")
        else:
            print(f"ℹ️ Se procesaron todas las {total_filas_disponibles} filas disponibles")
            
    except Exception as e:
        print(f"❌ Error general procesando filas DEOX: {str(e)}")
    
    return filas_deox

def extraer_texto_celda(celda):
    """Extrae el texto de una celda de DEOX"""
    try:
        span = celda.find_element(By.TAG_NAME, "span")
        return span.text.strip()
    except:
        return celda.text.strip()

def extraer_fecha_de_datetime(fecha_hora_str):
    """Extrae solo la fecha de un string que contiene fecha y hora"""
    try:
        if fecha_hora_str and len(fecha_hora_str) >= 10:
            fecha_parte = fecha_hora_str[:10]
            if re.match(r'\d{2}/\d{2}/\d{4}', fecha_parte):
                return fecha_parte
        return None
    except Exception as e:
        print(f"   ⚠️ Error extrayendo fecha de '{fecha_hora_str}': {e}")
        return None

def extraer_expediente_y_causa_deox(expediente_causa_str):
    """Extrae el expediente y la causa de un string combinado de DEOX"""
    expediente = "No disponible"
    causa = "No disponible"
    
    try:
        if expediente_causa_str:
            # Buscar el patrón del expediente al inicio
            exp_match = re.search(r'^([A-Z]{2,5}\s*\d{4,7}/\d{4})', expediente_causa_str)
            if exp_match:
                expediente = exp_match.group(1).strip()
                
                # La causa es todo lo que viene después del expediente y el guión
                causa_match = re.search(r'^[A-Z]{2,5}\s*\d{4,7}/\d{4}\s*-\s*(.+)', expediente_causa_str)
                if causa_match:
                    causa = causa_match.group(1).strip()
                else:
                    # Si no hay guión, tomar todo después del expediente
                    causa_match = re.search(r'^[A-Z]{2,5}\s*\d{4,7}/\d{4}\s+(.+)', expediente_causa_str)
                    if causa_match:
                        causa = causa_match.group(1).strip()
            
            print(f"       🔍 Extracción DEOX:")
            print(f"           Original: {expediente_causa_str[:100]}...")
            print(f"           Expediente: {expediente}")
            print(f"           Causa: {causa[:80]}...")
            
    except Exception as e:
        print(f"   ⚠️ Error extrayendo expediente y causa: {e}")
    
    return expediente, causa

def extraer_datos_mejorado(driver, fila, numero_fila, aria_label):
    """Extracción de datos de notificaciones via JavaScript DOM"""
    expediente = causa = juzgado = codigo = "No disponible"

    print(f"   🔍 Extrayendo datos de fila {numero_fila}...")

    try:
        data = driver.execute_script("""
            var el = arguments[0];
            var result = {codigo:'', expediente:'', causa:'', juzgado:''};
            
            var p = el.querySelector('p');
            if (p) result.codigo = p.textContent.trim();
            
            var td = null;
            if (p) {
                td = p.closest('td');
            } else {
                td = el.querySelector("td[data-key*='descripcion']") || el.querySelector("td[data-key*='Descripcion']");
            }
            if (!td) td = el;
            
            var allSpans = td.querySelectorAll('span');
            var spanTexts = [];
            for (var i=0; i<allSpans.length; i++) {
                var txt = allSpans[i].textContent.trim();
                // Ignorar guiones sueltos o vacíos
                if (txt && txt !== "-") {
                    spanTexts.push(txt);
                }
            }
            
            // Asignar en base al contenido
            for (var i=0; i<spanTexts.length; i++) {
                var txt = spanTexts[i];
                if (txt.match(/^[A-Z]{2,5}\\s*\\d{1,9}\\/\\d{4}/)) {
                    result.expediente = txt;
                } else if (txt.includes("JUZGADO") || txt.includes("SECRETARIA") || txt.includes("CAMARA") || txt.includes("TRIBUNAL") || txt.includes("SALA")) {
                    result.juzgado = txt;
                } else if (txt.includes(" c/ ") || txt.includes(" s/ ") || txt.length > 15) {
                    // Si no es expediente ni juzgado, y es largo, asumimos que es causa
                    if (!result.causa) {
                        result.causa = txt;
                    }
                }
            }
            
            return result;
        """, fila)

        if data.get('codigo'):
            codigo = data['codigo']
            print(f"   ✅ Código extraído: {codigo}")
        if data.get('expediente'):
            expediente = data['expediente']
            print(f"   ✅ Expediente extraído: {expediente}")
        if data.get('causa'):
            causa = data['causa']
            print(f"   ✅ Causa extraída: {causa}")
        if data.get('juzgado'):
            juzgado = data['juzgado']
            print(f"   ✅ Juzgado extraído: {juzgado}")

    except Exception as e:
        print(f"   ❌ Error en extracción JS: {e}")

    # Fallback aria-label para expediente
    if expediente == "No disponible" and aria_label:
        exp_match = re.search(r'expediente\s+([A-Z]{2,5}\s*\d{4,7}/\d{4}(?:/[^,]*)?)', aria_label, re.IGNORECASE)
        if exp_match:
            expediente = exp_match.group(1)
            print(f"   ✅ Expediente extraído del aria-label: {expediente}")
            
    # Fallback aria-label para causa
    if causa == "No disponible" and aria_label:
        causa_match = re.search(r'carátula\s+([^,]+)', aria_label, re.IGNORECASE)
        if causa_match:
            causa = causa_match.group(1).strip()
            print(f"   ✅ Causa extraída del aria-label: {causa}")

    # Fallback texto de fila completo para causa y juzgado si JS falló
    if causa == "No disponible" or juzgado == "No disponible":
        try:
            texto_fila = fila.text
            lineas = [line.strip() for line in texto_fila.split('\n') if line.strip()]
            if codigo in lineas:
                idx = lineas.index(codigo)
                # La estructura suele ser: [codigo, expediente - causa, juzgado]
                if idx + 1 < len(lineas):
                    if "-" in lineas[idx + 1]:
                        partes = lineas[idx + 1].split("-", 1)
                        if expediente == "No disponible":
                            expediente = partes[0].strip()
                        if causa == "No disponible":
                            causa = partes[1].strip()
                            print(f"   ✅ Causa extraída del texto de fila: {causa[:30]}...")
                    elif causa == "No disponible":
                        causa = lineas[idx + 1]
                        print(f"   ✅ Causa extraída del texto de fila: {causa[:30]}...")
                        
                if idx + 2 < len(lineas) and juzgado == "No disponible":
                    juzgado = lineas[idx + 2]
                    print(f"   ✅ Juzgado extraído del texto de fila: {juzgado[:30]}...")
        except Exception as e:
            print(f"   ⚠️ Fallback de texto falló: {e}")

    print(f"   📊 RESULTADO FINAL FILA {numero_fila}:")
    print(f"       Código: {codigo}")
    print(f"       Expediente: {expediente}")
    print(f"       Causa: {(causa[:80] + '...') if causa != 'No disponible' else causa}")
    print(f"       Juzgado: {(juzgado[:60] + '...') if juzgado != 'No disponible' else juzgado}")

    return expediente, causa, juzgado, codigo

def ir_siguiente_pagina(driver):
    """Intenta ir a la siguiente página con una estrategia más robusta."""
    print("   🔄 Buscando botón para ir a la siguiente página (nueva estrategia)...")

    # Obtener el contenido actual de la página para verificar cambios
    old_page_source = driver.page_source
    
    try:
        # Estrategia 1: Buscar el botón "Siguiente" por su título o texto
        print("   → Estrategia 1: Buscando botón 'Siguiente' por título/texto...")
        next_button_xpath = "//a[contains(@title, 'Siguiente') or .//span[@title='Siguiente'] or contains(text(), 'Siguiente')]"
        next_buttons = driver.find_elements(By.XPATH, next_button_xpath)
        
        clicked = False
        for btn in next_buttons:
            if btn.is_displayed() and btn.is_enabled() and "disabled" not in btn.get_attribute("class"):
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable(btn))
                    btn.click()
                    print("   ✅ Click en botón 'Siguiente' exitoso.")
                    clicked = True
                    break
                except Exception as e:
                    print(f"   ⚠️ Falló click en botón 'Siguiente': {e}")
        
        if clicked:
            print("   ⏳ Esperando carga de nueva página después del click...")
            time.sleep(5) # Dar tiempo para que cargue
            # Verificar cambio de página comparando el page_source
            if driver.page_source != old_page_source:
                print("   ✅ Contenido de la página cambió. Navegación exitosa.")
                return True
            else:
                print("   ❌ Contenido de la página NO cambió. Reintentando.")
                # No retornar False aquí, intentar la siguiente estrategia
        
        # Estrategia 2: Buscar el número de página actual y hacer clic en el siguiente número
        print("   → Estrategia 2: Buscando número de página siguiente...")
        current_page_element = None
        try:
            current_page_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'current-page') or contains(@class, 'active-page')]"))
            )
            current_page_num = int(current_page_element.text.strip())
            next_page_num = current_page_num + 1
            print(f"   📍 Página actual: {current_page_num}, Intentando ir a: {next_page_num}")

            # Intentar encontrar el enlace para el siguiente número de página
            next_page_link_xpath = f"//a[.//span[text()='{next_page_num}'] or @data-dt-idx='{next_page_num}']"
            next_page_link = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, next_page_link_xpath))
            )
            
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_page_link)
            next_page_link.click()
            print(f"   ✅ Click en número de página {next_page_num} exitoso.")
            
            print("   ⏳ Esperando carga de nueva página después del click...")
            time.sleep(5) # Dar tiempo para que cargue
            if driver.page_source != old_page_source:
                print("   ✅ Contenido de la página cambió. Navegación exitosa.")
                return True
            else:
                print("   ❌ Contenido de la página NO cambió. Reintentando.")
                return False

        except Exception as e:
            print(f"   ⚠️ Estrategia 2 falló: {e}")
            
        print("   ❌ No se pudo navegar a la siguiente página con ninguna estrategia.")
        return False

    except Exception as e:
        print(f"   ❌ Error general en ir_siguiente_pagina: {str(e)}")
        return False

def guardar_todos_en_excel(todos_los_datos, fecha_objetivo):
    """Guarda todos los datos en Excel"""
    try:
        nombre_archivo = "expedientes.xlsx"
        nombre_hoja = "Expedientes"
        
        if not todos_los_datos:
            print(f"\n💾 No hay datos nuevos para agregar")
            return
        
        # Preparar DataFrame con los nuevos datos
        df_nuevos = pd.DataFrame(todos_los_datos)
        
        # Agregar columna de código si no existe
        if 'codigo' not in df_nuevos.columns:
            df_nuevos['codigo'] = 'N/A'
        
        # Agregar columnas de número y año si no existen
        if 'numero_expediente' not in df_nuevos.columns:
            df_nuevos['numero_expediente'] = ''
        if 'ano_expediente' not in df_nuevos.columns:
            df_nuevos['ano_expediente'] = ''
        
        # Reordenar columnas
        df_nuevos = df_nuevos[['expediente', 'numero_expediente', 'ano_expediente', 'juzgado', 'causa', 'estado', 'fecha', 'fuente', 'codigo']]
        df_nuevos.columns = ['Expediente', 'Número', 'Año', 'Juzgado', 'Causa', 'Estado', 'Fecha', 'Fuente', 'Código']
        
        # Agregar timestamp
        timestamp_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df_nuevos['Agregado'] = timestamp_actual
        
        if os.path.exists(nombre_archivo):
            print(f"\n💾 Archivo '{nombre_archivo}' existe. Agregando datos...")
            
            try:
                df_existentes = pd.read_excel(nombre_archivo, sheet_name=nombre_hoja)
                print(f"   📊 Registros existentes: {len(df_existentes)}")
                
                # Verificar duplicados
                duplicados_encontrados = 0
                df_sin_duplicados = []
                
                for _, fila_nueva in df_nuevos.iterrows():
                    # Normalizar fechas para comparar
                    exp_nuevo = fila_nueva['Expediente']
                    fecha_nueva = normalizar_fecha_para_excel(fila_nueva['Fecha'])
                    duplicado = df_existentes[
                        (df_existentes['Expediente'] == exp_nuevo) &
                        (df_existentes['Fecha'].apply(lambda f: normalizar_fecha_para_excel(str(f))) == fecha_nueva)
                    ]
                    if duplicado.empty:
                        df_sin_duplicados.append(fila_nueva)
                    else:
                        duplicados_encontrados += 1
                        print(f"   ⚠️ Duplicado: {exp_nuevo} - {fecha_nueva}")
                
                if df_sin_duplicados:
                    df_sin_duplicados = pd.DataFrame(df_sin_duplicados)
                    df_final = pd.concat([df_existentes, df_sin_duplicados], ignore_index=True)
                    
                    with pd.ExcelWriter(nombre_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_final.to_excel(writer, sheet_name=nombre_hoja, index=False)
                    
                    print(f"   ✅ {len(df_sin_duplicados)} registros nuevos agregados")
                    print(f"   ⚠️ {duplicados_encontrados} duplicados omitidos")
                    print(f"   📊 Total registros: {len(df_final)}")
                else:
                    print(f"   ℹ️ No hay registros nuevos (todos son duplicados)")
                    
            except ValueError as e:
                if "not found" in str(e):
                    print(f"   ℹ️ Hoja '{nombre_hoja}' no existe, creándola...")
                    with pd.ExcelWriter(nombre_archivo, engine='openpyxl', mode='a') as writer:
                        df_nuevos.to_excel(writer, sheet_name=nombre_hoja, index=False)
                    print(f"   ✅ Hoja creada con {len(df_nuevos)} registros")
                else:
                    raise e
        else:
            print(f"\n💾 Creando nuevo archivo '{nombre_archivo}'...")
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
                df_nuevos.to_excel(writer, sheet_name=nombre_hoja, index=False)
            print(f"   ✅ Hoja creada con {len(df_nuevos)} registros")
        
        # Contar por fuente
        consultas_count = len([d for d in todos_los_datos if d['fuente'] == 'Consultas'])
        notificaciones_count = len([d for d in todos_los_datos if d['fuente'] == 'Notificaciones'])
        deox_count = len([d for d in todos_los_datos if d['fuente'] == 'DEOX'])
        
        print(f"\n📊 RESUMEN DE ESTA EJECUCIÓN:")
        print(f"   - Consultas: {consultas_count}")
        print(f"   - Notificaciones: {notificaciones_count}")
        print(f"   - DEOX: {deox_count}")
        print(f"   - Total: {len(todos_los_datos)}")
        print(f"📂 Archivo: {os.path.abspath(nombre_archivo)}")
        
        # Aplicar formato
        aplicar_formato_excel(nombre_archivo, nombre_hoja)
        
    except Exception as e:
        print(f"\n❌ Error al guardar en Excel: {str(e)}")

def aplicar_formato_excel(nombre_archivo, nombre_hoja):
    """Aplica formato básico al archivo Excel"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = load_workbook(nombre_archivo)
        
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Aplicar formato a encabezados
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 20,  # Expediente
                'B': 12,  # Número
                'C': 8,   # Año
                'D': 30,  # Juzgado
                'E': 50,  # Causa
                'F': 15,  # Estado
                'G': 12,  # Fecha
                'H': 15,  # Fuente
                'I': 15,  # Código
                'J': 20   # Agregado
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(nombre_archivo)
            print(f"   ✨ Formato aplicado exitosamente")
            
    except Exception as e:
        print(f"   ⚠️ Error aplicando formato: {str(e)}")

def mostrar_resumen_final(filas_consultas, filas_notificaciones, filas_deox, fecha_objetivo):
    """Muestra un resumen final de la ejecución"""
    print("\n" + "="*80)
    print("🎯 RESUMEN FINAL DE LA EJECUCIÓN")
    print("="*80)
    print(f"📅 Fecha objetivo: {fecha_objetivo}")
    print(f"🔍 Consultas encontradas: {len(filas_consultas)}")
    print(f"🔔 Notificaciones encontradas: {len(filas_notificaciones)}")
    print(f"📋 DEOX encontrados: {len(filas_deox)}")
    print(f"📊 TOTAL DE REGISTROS: {len(filas_consultas) + len(filas_notificaciones) + len(filas_deox)}")
    print("="*80)
    
    if filas_consultas:
        print("\n🔍 EXPEDIENTES DE CONSULTAS:")
        for i, fila in enumerate(filas_consultas, 1):
            print(f"   {i}. {fila['expediente']} - {fila['causa'][:60]}...")
    
    if filas_notificaciones:
        print("\n🔔 EXPEDIENTES DE NOTIFICACIONES:")
        for i, fila in enumerate(filas_notificaciones, 1):
            print(f"   {i}. {fila['expediente']} - {fila['causa'][:60]}...")
    
    if filas_deox:
        print("\n📋 EXPEDIENTES DE DEOX:")
        for i, fila in enumerate(filas_deox, 1):
            print(f"   {i}. {fila['expediente']} - {fila['causa'][:60]}...")
    
    print("\n✅ Proceso completado exitosamente")
    print("📂 Los datos han sido guardados en 'expedientes.xlsx'")
    print("="*80)

def guardar_datos_en_excel_analizados(df_resultado, archivo_salida):
    """
    NUEVA FUNCIÓN: Guarda los datos analizados en Excel con columnas separadas
    """
    try:
        # Definir todas las columnas que queremos (incluyendo las originales)
        columnas_originales = list(df_resultado.columns)
        
        # Columnas de análisis que queremos asegurar
        columnas_analisis = [
            'Expediente_Numero', 'Jurisdiccion_Normalizada', 'Juzgado_Normalizado', 
            'Secretaria', 'Situacion_Actual', 'Tipo_Proceso', 'Actor_Principal', 
            'Demandado_Principal', 'Objeto_Demanda', 'Letrado_Actor', 
            'Matricula_Letrado', 'CUIT_CUIL_Letrado', 'Observaciones', 'Estado_Analisis'
        ]
        
        # Asegurar que el DataFrame tenga todas las columnas de análisis
        for col in columnas_analisis:
            if col not in df_resultado.columns:
                df_resultado[col] = 'No disponible'
                print(f"   ➕ Agregando columna: {col}")
        
        # Mostrar información de debug
        print(f"📊 Columnas en el DataFrame:")
        for col in df_resultado.columns:
            valores_no_vacios = df_resultado[col].notna().sum()
            print(f"   {col}: {valores_no_vacios} valores no vacíos")
        
        # Guardar en Excel
        df_resultado.to_excel(archivo_salida, index=False)
        print(f"💾 Datos guardados en '{archivo_salida}' con {len(df_resultado)} filas y {len(df_resultado.columns)} columnas")
        
        return True
        
    except Exception as e:
        print(f"❌ Error guardando datos en Excel: {e}")
        traceback.print_exc()
        return False

if __name__ == "__main__":
    # Ejemplo de uso para extracción
    fecha = "8/07/2025"
    paginas = 3
    usuario = "tu_usuario"
    password = "tu_password"
    
    print("🚀 Iniciando filtrado por fecha...")
    filtrar_por_fecha(fecha, paginas, usuario, password, headless=False, filas_deox=10)
    
    # Ejemplo de uso para análisis individual
    print("🤖 Iniciando análisis individual de expedientes...")
    analizar_expedientes_individuales(usuario, password, headless=False)
